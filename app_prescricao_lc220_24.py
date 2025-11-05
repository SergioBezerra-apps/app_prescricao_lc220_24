# app_prescricao_lc220_24.py
import streamlit as st
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from datetime import date as _date_for_prevcheck
import pandas as pd
from io import BytesIO
import zipfile
import re

# ======================================================================================
# Config & estilo
# ======================================================================================
st.set_page_config(page_title="Prescrição — LC-RJ 63/1990 (art. 5º-A)", layout="wide")
st.markdown("<style>.block-container {max-width:980px; padding-left:12px; padding-right:12px;}</style>", unsafe_allow_html=True)

# ======================================================================================
# Utils de data / formatação (PT-BR)
# ======================================================================================
def fmt_br(d: date | None) -> str:
    return d.strftime("%d/%m/%Y") if isinstance(d, date) else "—"

def di(label: str, default: date | None = None, key: str | None = None, help: str | None = None) -> date:
    """date_input sem 'trava dos 10 anos'; aceita qualquer data válida do widget."""
    return st.date_input(label, value=(default or date.today()), key=key, help=help)

# ======================================================================================
# Utilitários para gerar DOCX (Roteiro Oficial) — sem dependências externas
# ======================================================================================
def _xml_escape(s: str) -> str:
    return (s.replace("&", "&amp;")
             .replace("<", "&lt;")
             .replace(">", "&gt;")
             .replace('"', "&quot;")
             .replace("'", "&apos;"))

def _wpara(text: str, bold=False, size=24) -> str:
    t = _xml_escape(text)
    if bold:
        return (f"<w:p><w:r><w:rPr><w:b/><w:sz w:val='{size}'/></w:rPr>"
                f"<w:t xml:space='preserve'>{t}</w:t></w:r></w:p>")
    return f"<w:p><w:r><w:t xml:space='preserve'>{t}</w:t></w:r></w:p>"

def build_roteiro_docx_bytes(with_videos: bool = True) -> bytes:
    sections: list[tuple[str, bool]] = []
    add = lambda txt, h=False: sections.append((txt, h))

    # Título
    add("ROTEIRO OFICIAL — Calculadora de Prescrição (LC-RJ 63/1990, art. 5º-A)", True)

    # 1) Finalidade
    add("1) Finalidade", True)
    add("Padronizar a aplicação do art. 5º-A (LCE 63/1990) no TCE-RJ, seguindo a chave intertemporal consolidada "
        "em votos paradigmáticos (p.ex., 114.199-4/2024 e 227.877-1/2014): fatos ≥ 18/07/2021 → novo regime (5 anos do "
        "fato/cessação); fatos < 18/07/2021 → teste pré-lei (quinquênio da ciência até 18/07/2024) e, se não consumado, "
        "transição bienal (18/07/2024 → 18/07/2026), ainda que a ciência seja posterior.")

    # 2) Chave intertemporal (executiva)
    add("2) Chave intertemporal — visão executiva", True)
    add("• Fatos ≥ 18/07/2021 → Novo regime (5 anos do fato/cessação).")
    add("• Fatos < 18/07/2021 → Faça o TESTE PRÉ-LEI: projete 5 anos da ciência (em regra, autuação institucional) com marcos até 18/07/2024. "
        "Se consumou até 18/07/2024 → prescrição antes da lei. Se NÃO consumou → Transição (18/07/2024 → 18/07/2026), "
        "independentemente de a ciência ocorrer depois.")
    add("• Decisão administrativa transitada até 18/07/2024 → fora do alcance da LCE 220/2024.")

    # 3) Marcos interruptivos
    add("3) Marcos interruptivos (§ 3º)", True)
    add("• Teste pré-lei: considere apenas marcos entre a ciência e 18/07/2024 (reiniciam o quinquênio do regime anterior).")
    add("• Transição: só contam marcos a partir de 18/07/2024 (reiniciam o bienal).")
    add("• Novo regime: marcos a partir do termo material (fato/cessação).")
    add("• Qualificação: chamamento qualificado tem efeito subjetivo (por gestor) e retroage à decisão que o determinou. "
        "Simples protocolo não interrompe.")

    # 4) Intercorrente e prazo penal
    add("4) Intercorrente e prazo penal", True)
    add("• Intercorrente (§ 1º): paralisação superior a 3 anos sem julgamento ou despacho útil.")
    add("• Prazo penal (§ 2º): prevalece sobre o administrativo quando cabível.")

    # 5) Passo a passo de uso
    add("5) Passo a passo de uso", True)
    add("1. Defina natureza, conduta, termo material (fato/cessação) e informe a autuação e a ciência (se diversa).")
    add("2. Lance marcos gerais (objetivos, valem para todos) e os chamamentos qualificados por gestor (efeito subjetivo).")
    add("3. O app sugere o enquadramento global: novo regime / transição / prescrição antes da lei / fora do alcance — ajuste se necessário.")
    add("4. Se necessário, calcule intercorrente (último ato × ato subsequente/hoje).")
    add("5. Analise os cartões por gestor e exporte o Excel com Resumo + abas auxiliares (parâmetros, marcos, dicionário e abas por gestor).")

    # 6) Índice e explicações dos vídeos (sem legenda/sem áudio)
    if with_videos:
        add("6) ÍNDICE DE VÍDEOS (arquivos .mp4 sem áudio/sem legenda)", True)

        def vid(title, objetivo, inputs, resultado):
            add(title, True)
            add(f"Objetivo: {objetivo}")
            add(f"Inputs-chave: {inputs}")
            add(f"Resultado esperado: {resultado}")

        vid("01_Novo_Regime_FatoRecente.mp4",
            "Fato ≥ 18/07/2021 com contagem quinquenal do fato/cessação.",
            "Punitiva; Ato 03/11/2021; Autuação 12/12/2024; sem marcos; sem intercorrente.",
            "Enquadramento: “Novo regime (art. 5º-A)”; prazo final 03/11/2026.")
        vid("02_Transicao_CienciaPosterior.mp4",
            "Fato anterior a 18/07/2021 com ciência posterior: aplica transição bienal.",
            "Punitiva; Ato 15/06/2016; Ciência 12/12/2024; sem marcos.",
            "“Transição 2 anos (LC 220/24)”; vence 18/07/2026.")
        vid("03_Prescricao_AntesDaLei.mp4",
            "Reconhecer prescrição pré-lei pelo quinquênio da ciência.",
            "Ato 10/05/2015; Ciência 10/06/2017; sem marcos até 18/07/2024.",
            "“Prescrição reconhecida (regime anterior)”.")
        vid("04_Transicao_MarcoGeral.mp4",
            "Mostrar reinício do bienal por ato inequívoco de apuração pós-lei.",
            "Ato 20/02/2017; Ciência 01/08/2024; Marco geral 10/09/2025.",
            "Novo vencimento 10/09/2027.")
        vid("05_Transicao_Chamamento_Subjetivo.mp4",
            "Efeito subjetivo do chamamento qualificado (multi-gestores).",
            "Fato 2016; Ciência 2024; Gestor A com chamamento 20/06/2026 (decisão em 05/05/2026); Gestor B sem chamamento.",
            "Gestor A vence 05/05/2028 (retroação); Gestor B vence 18/07/2026.")
        vid("06_Intercorrente.mp4",
            "Prescrição intercorrente (> 3 anos) durante a tramitação.",
            "Novo regime; último ato 01/08/2021; ato subsequente 05/09/2024.",
            "“Prescrição intercorrente”.")
        vid("07_Continuada_Cessacao.mp4",
            "Conduta continuada (termo na cessação).",
            "Cessação 31/12/2022; sem marcos; sem intercorrente.",
            "Novo regime; prazo final 31/12/2027.")
        vid("08_Ressarcitoria_UltimaMedicao.mp4",
            "Ressarcitória (analogia) com base “última medição/pagamento”.",
            "Última medição 30/03/2019; ciência 2024; sem marcos.",
            "Transição; vence 18/07/2026 (salvo marcos pós-lei).")
        vid("09_PrazoPenal.mp4",
            "Prevalência do prazo penal (§ 2º).",
            "Ato 10/10/2022; “Fato também é crime: Sim”; Prazo penal 8 anos.",
            "Base = penal (8 anos); vencimento 10/10/2030.")
        vid("10_Ciencia_Apos_18072026.mp4",
            "Ciência somente após 18/07/2026 em caso de transição (sem marcos).",
            "Fato 2017; Ciência 01/08/2026; sem marcos.",
            "Prescrição consumada em 18/07/2026 (ciência tardia não reabre).")
        vid("11_Multigestores_ExportacaoExcel.mp4",
            "Preencher vários gestores e exportar o Excel completo.",
            "Fato 2016; Marco geral 01/03/2025; Chamamento apenas do Gestor B em 15/05/2025.",
            "Planilha com Resumo + abas auxiliares; prazos distintos por gestor.")

    # Montagem do DOCX mínimo (WordprocessingML)
    body_xml = "".join(_wpara(t, bold=h, size=28 if h else 24) for t, h in sections)
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:wpc="http://schemas.microsoft.com/office/2010/wordprocessingCanvas" '
        'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" '
        'xmlns:o="urn:schemas-microsoft-com:office:office" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
        'xmlns:m="http://schemas.openxmlformats.org/officeDocument/2006/math" '
        'xmlns:v="urn:schemas-microsoft-com:vml" '
        'xmlns:wp14="http://schemas.microsoft.com/office/2010/wordprocessingDrawing" '
        'xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing" '
        'xmlns:w10="urn:schemas-microsoft-com:office:word" '
        'xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main" '
        'xmlns:w14="http://schemas.microsoft.com/office/2010/wordml" '
        'xmlns:wpg="http://schemas.microsoft.com/office/2010/wordprocessingGroup" '
        'xmlns:wpi="http://schemas.microsoft.com/office/2010/wordprocessingInk" '
        'xmlns:wne="http://schemas.microsoft.com/office/2006/wordml" '
        'xmlns:wps="http://schemas.microsoft.com/office/2010/wordprocessingShape" mc:Ignorable="w14 wp14">'
        '<w:body>' +
        body_xml +
        '<w:sectPr><w:pgSz w:w="12240" w:h="15840"/>'
        '<w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440" w:header="708" w:footer="708" w:gutter="0"/></w:sectPr>'
        '</w:body></w:document>'
    )
    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '</Types>'
    )
    rels_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>'
        '</Relationships>'
    )
    word_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<Relationships xmlns="http://schemas.microsoft.com/office/2006/relationships"/>'
    )

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        z.writestr('[Content_Types].xml', content_types_xml)
        z.writestr('_rels/.rels', rels_xml)
        z.writestr('word/document.xml', document_xml)
        z.writestr('word/_rels/document.xml.rels', word_rels_xml)
    return buf.getvalue()

# ======================================================================================
# Helpers Excel
# ======================================================================================
def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[:\\/?*\[\]]', '_', name).strip()
    return name[:31] if len(name) > 31 else name

def make_excel_bytes_expanded(rows_resumo: list[dict],
                              rows_marcos_gerais: list[dict],
                              rows_marcos_subj: list[dict],
                              parametros: dict,
                              por_gestor_details: dict) -> bytes:
    """Gera .xlsx com fallback: usa 'xlsxwriter' se disponível (format condicional), senão 'openpyxl'."""
    engine = "openpyxl"
    try:
        import xlsxwriter  # noqa: F401
        engine = "xlsxwriter"
    except Exception:
        engine = "openpyxl"

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine=engine, datetime_format="dd/mm/yyyy", date_format="dd/mm/yyyy") as writer:
        # Resumo
        df_resumo = pd.DataFrame(rows_resumo) if rows_resumo else pd.DataFrame(columns=[
            "Gestor","Enquadramento","Situação","Base","Termo material","Ciência (TCE-RJ)",
            "Termo inicial efetivo","Data atual de prescrição","Interrupções consideradas"
        ])
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        ws_resumo = writer.sheets["Resumo"]

        # Marcos_Gerais
        df_g = pd.DataFrame(rows_marcos_gerais) if rows_marcos_gerais else pd.DataFrame(columns=["marco_geral_data"])
        df_g.to_excel(writer, sheet_name="Marcos_Gerais", index=False)
        ws_g = writer.sheets["Marcos_Gerais"]

        # Marcos_Subjetivos
        df_s = pd.DataFrame(rows_marcos_subj) if rows_marcos_subj else pd.DataFrame(columns=["gestor","chamamento_data"])
        df_s.to_excel(writer, sheet_name="Marcos_Subjetivos", index=False)
        ws_s = writer.sheets["Marcos_Subjetivos"]

        # Parametros_do_Caso
        p_rows = [(k, v) for k, v in parametros.items()]
        df_p = pd.DataFrame(p_rows, columns=["parametro", "valor"])
        df_p.to_excel(writer, sheet_name="Parametros_do_Caso", index=False)
        ws_p = writer.sheets["Parametros_do_Caso"]

        # Dicionario
        dic_data = [
            ("Gestor", "Nome do gestor (uma linha por gestor)."),
            ("Enquadramento", "Novo regime / Transição 2 anos / Prescrição antes da lei / Fora do alcance."),
            ("Situação", "Não prescrito / Prescrição consumada / Prescrição intercorrente / Prescrição reconhecida (regime anterior)."),
            ("Base", "quinquenal / penal (X anos) / bienal (transição)."),
            ("Termo material", "Data do fato/cessação (ou base motivada — ressarcitória)."),
            ("Ciência (TCE-RJ)", "Data de ciência considerada (em regra, autuação)."),
            ("Termo inicial efetivo", "Data usada no cálculo (depende do enquadramento)."),
            ("Data atual de prescrição", "Data-alvo projetada após interrupções."),
            ("Interrupções consideradas", "Marcos gerais pós-lei e chamamentos por gestor."),
            ("marco_geral_data", "Ato inequívoco de apuração/decisão recorrível/tentativa conciliatória (vale para todos)."),
            ("chamamento_data", "Decisão do chamamento qualificado (efeito subjetivo)."),
            ("parametro/valor", "Parâmetros do caso — contexto global."),
        ]
        df_dic = pd.DataFrame(dic_data, columns=["coluna", "descrição"])
        df_dic.to_excel(writer, sheet_name="Dicionario", index=False)
        ws_d = writer.sheets["Dicionario"]

        # Abas individuais por gestor
        for g, detail in por_gestor_details.items():
            sheet = sanitize_sheet_name(f"G - {g}")
            df_det = pd.DataFrame(detail["linhas"])
            if df_det.empty:
                df_det = pd.DataFrame(columns=["campo", "valor"])
            df_det.to_excel(writer, sheet_name=sheet, index=False)
            ws_x = writer.sheets[sheet]

        # Formatação e freeze
        if engine == "xlsxwriter":
            wb = writer.book
            # Resumo
            widths = [26, 24, 26, 18, 18, 18, 22, 22, 60]
            for i, w in enumerate(widths):
                ws_resumo.set_column(i, i, w)
            ws_resumo.freeze_panes(1, 0)
            red_fmt = wb.add_format({"font_color": "#D93025"})
            green_fmt = wb.add_format({"font_color": "#1E8E3E"})
            blue_fmt = wb.add_format({"font_color": "#1A73E8"})
            last_row = len(df_resumo) + 1
            ws_resumo.conditional_format(f"C2:C{last_row}", {"type": "text", "criteria": "containing", "value": "Prescrição consumada", "format": red_fmt})
            ws_resumo.conditional_format(f"C2:C{last_row}", {"type": "text", "criteria": "containing", "value": "intercorrente", "format": red_fmt})
            ws_resumo.conditional_format(f"C2:C{last_row}", {"type": "text", "criteria": "containing", "value": "Não prescrito", "format": green_fmt})
            ws_resumo.conditional_format(f"C2:C{last_row}", {"type": "no_blanks", "format": blue_fmt})

            # Demais abas
            ws_g.set_column("A:A", 18); ws_g.freeze_panes(1, 0)
            ws_s.set_column("A:A", 28); ws_s.set_column("B:B", 18); ws_s.freeze_panes(1, 0)
            ws_p.set_column("A:A", 36); ws_p.set_column("B:B", 60); ws_p.freeze_panes(1, 0)
            ws_d.set_column("A:A", 30); ws_d.set_column("B:B", 90); ws_d.freeze_panes(1, 0)
            for g in por_gestor_details.keys():
                sheet = sanitize_sheet_name(f"G - {g}")
                ws_x = writer.sheets[sheet]
                ws_x.set_column("A:A", 34)
                ws_x.set_column("B:B", 70)
                ws_x.freeze_panes(1, 0)
        else:
            from openpyxl.utils import get_column_letter
            # Resumo
            widths = [26, 24, 26, 18, 18, 18, 22, 22, 60]
            for idx, w in enumerate(widths, start=1):
                ws_resumo.column_dimensions[get_column_letter(idx)].width = w
            ws_resumo.freeze_panes = "A2"
            # Demais abas
            ws_g.column_dimensions[get_column_letter(1)].width = 18; ws_g.freeze_panes = "A2"
            ws_s.column_dimensions[get_column_letter(1)].width = 28
            ws_s.column_dimensions[get_column_letter(2)].width = 18
            ws_s.freeze_panes = "A2"
            ws_p.column_dimensions[get_column_letter(1)].width = 36
            ws_p.column_dimensions[get_column_letter(2)].width = 60
            ws_p.freeze_panes = "A2"
            ws_d.column_dimensions[get_column_letter(1)].width = 30
            ws_d.column_dimensions[get_column_letter(2)].width = 90
            ws_d.freeze_panes = "A2"
            for g in por_gestor_details.keys():
                sheet = sanitize_sheet_name(f"G - {g}")
                ws_x = writer.sheets[sheet]
                ws_x.column_dimensions[get_column_letter(1)].width = 34
                ws_x.column_dimensions[get_column_letter(2)].width = 70
                ws_x.freeze_panes = "A2"

    return buf.getvalue()

# ======================================================================================
# Cabeçalho + Roteiro (DOCX)
# ======================================================================================
st.title("Calculadora de Prescrição — LC-RJ 63/1990 (art. 5º-A)")
st.caption("Ferramenta de apoio. Ajuste as premissas ao caso concreto e registre a motivação no parecer.")

with st.expander("📘 Roteiro Oficial — ver/baixar", expanded=False):
    st.markdown("O Roteiro Oficial consolida regras, chave intertemporal e exemplos de uso (vídeos).")
    roteiro_bytes = build_roteiro_docx_bytes(with_videos=True)
    st.download_button(
        "⬇️ Baixar Roteiro Oficial (DOCX)",
        data=roteiro_bytes,
        file_name="Roteiro_Oficial_Calculadora_Prescricao.docx",
        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        use_container_width=True
    )

# ======================================================================================
# 1) Natureza e dados básicos
# ======================================================================================
colA, colB, colC = st.columns([1.2, 1, 1])
with colA:
    natureza = st.selectbox(
        "Natureza da pretensão",
        ["Punitiva", "Ressarcitória (analogia)"],
        help=("Selecione Punitiva (ex.: multa) ou Ressarcitória (analogia). "
              "A LCE 220/2024 (art. 5º-A) rege a prescrição no TCE-RJ; a ressarcitória segue por analogia consolidada."),
    )
with colB:
    conduta = st.selectbox(
        "Tipo de conduta",
        ["Instantânea", "Continuada"],
        help="Instantânea: ato único. Continuada: efeitos que perduram (use a data de cessação).",
    )
with colC:
    data_autuacao = di(
        "Data de autuação no TCE-RJ",
        default=date.today(),
        help="Em regra, funciona como ciência institucional quando não houver prova de ciência diversa."
    )

data_ciencia = di(
    "Data de ciência pelo TCE-RJ (se diversa da autuação)",
    default=data_autuacao,
    help=("Para fatos anteriores a 18/07/2021, o teste pré-lei considera o quinquênio a partir da ciência "
          "(paradigma histórico: autuação institucional, salvo prova de ciência anterior)."),
)

# Termo material (fato/cessação ou base ressarcitória)
st.subheader("Termo material (fato/evento)")
if natureza == "Punitiva":
    data_ato = di(
        "Data do ato (ou da cessação, se continuada)",
        default=date.today(),
        help=("No novo regime (art. 5º-A), o termo é o fato/cessação. "
              "Essa data aciona a chave intertemporal: < 18/07/2021 (passivo antigo); ≥ 18/07/2021 (novo regime)."),
    )
    termo_material = data_ato
    termo_material_label = "Data do ato/cessação (punitiva)"
else:
    st.markdown("**Ressarcitória (analogia)** — motive a base do termo inicial no parecer.")
    base_ress = st.radio(
        "Base do termo (ressarcitória)",
        ["Evento danoso (data do dano)", "Última medição/pagamento (contratos)", "Cessação do dano (se continuada)"],
        help="A base escolhida deve ser fundamentada."
    )
    if base_ress == "Evento danoso (data do dano)":
        data_base = di("Data do evento danoso", default=date.today())
    elif base_ress == "Última medição/pagamento (contratos)":
        data_base = di("Data da última medição/pagamento ligada ao sobrepreço/irregularidade", default=date.today())
    else:
        data_base = di("Data de cessação do dano", default=date.today())
    termo_material = data_base
    termo_material_label = base_ress

colD, colE, colF = st.columns(3)
with colD:
    decisao_transitada_pre_lc = st.selectbox(
        "Decisão adm. transitada em julgado antes de 18/07/2024?",
        ["Não", "Sim"],
        help="Se 'Sim', a LCE 220/2024 não alcança (ato findo).",
    )
with colE:
    aplicar_prazo_penal = st.selectbox(
        "Fato também é crime? (aplica prazo penal)",
        ["Não", "Sim"],
        help="Se houver tipificação penal aplicável, prevalece o prazo penal (art. 5º-A, § 2º).",
    )
with colF:
    prazo_penal_anos = None
    if aplicar_prazo_penal == "Sim":
        prazo_penal_anos = st.number_input("Prazo penal (anos)", min_value=1, max_value=40, value=8, step=1)

# ======================================================================================
# 2) Marcos interruptivos — gerais x subjetivos
# ======================================================================================
st.subheader("Marcos interruptivos (§ 3º)")
st.caption(
    "Marcos gerais (objetivos, valem para todos): p.ex., determinação formal de auditoria/instauração de TCE/TOF, "
    "decisão condenatória recorrível, tentativa de conciliação. Simples protocolo não interrompe.\n"
    "Marcos subjetivos (por gestor): chamamento qualificado (efeito subjetivo; retroage à decisão que o determinou)."
)

# Marcos gerais
st.markdown("#### Marcos gerais (valem para todos)")
def _init_g_state():
    if "g_marco_count" not in st.session_state:
        st.session_state.g_marco_count = 1
    if "g_marco_dates" not in st.session_state:
        st.session_state.g_marco_dates = [None]
_init_g_state()

no_global_inter = st.checkbox("Não houve marco geral", value=False)
def _g_add():
    st.session_state.g_marco_count += 1
    st.session_state.g_marco_dates.append(None)
def _g_rem():
    if st.session_state.g_marco_count > 1:
        st.session_state.g_marco_count -= 1
        st.session_state.g_marco_dates = st.session_state.g_marco_dates[: st.session_state.g_marco_count]
def _g_clr():
    st.session_state.g_marco_count = 1
    st.session_state.g_marco_dates = [None]

g_interrupcoes: list[date] = []
if not no_global_inter:
    for i in range(st.session_state.g_marco_count):
        default_val = st.session_state.g_marco_dates[i] or date.today()
        picked = di(f"Data do marco geral #{i+1}", default=default_val, key=f"g_marco_{i}")
        st.session_state.g_marco_dates[i] = picked
    colA1, colA2, colA3 = st.columns(3)
    colA1.button("➕ Adicionar marco geral", use_container_width=True, on_click=_g_add)
    colA2.button("➖ Remover último", disabled=st.session_state.g_marco_count <= 1, use_container_width=True, on_click=_g_rem)
    colA3.button("🗑️ Limpar todos", use_container_width=True, on_click=_g_clr)
    g_interrupcoes = [d for d in st.session_state.g_marco_dates if isinstance(d, date)]
else:
    g_interrupcoes = []

st.markdown("---")

# Lista de gestores
st.markdown("#### Gestores (um por linha)")
gestores_text = st.text_area(
    "Nomes dos gestores",
    value="Gestor A\nGestor B",
    height=90,
    help="Indique um gestor por linha. Para cada gestor, informe os chamamentos qualificados (efeito subjetivo).",
)
gestores = [g.strip() for g in gestores_text.splitlines() if g.strip()]

# Chamamentos qualificados por gestor (efeito subjetivo)
st.markdown("#### Chamamentos qualificados por gestor (efeito subjetivo)")
if "gestor_marcos" not in st.session_state:
    st.session_state.gestor_marcos = {}  # nome -> [dates]
for g in gestores:
    if g not in st.session_state.gestor_marcos:
        st.session_state.gestor_marcos[g] = []

def _ensure_g_state(g):
    cnt_key = f"{g}__cnt"
    if cnt_key not in st.session_state:
        st.session_state[cnt_key] = 1
        st.session_state.gestor_marcos[g] = [None]
    return cnt_key

for g in gestores:
    with st.expander(f"Chamamentos qualificados — {g}", expanded=False):
        cnt_key = _ensure_g_state(g)
        no_subj = st.checkbox(f"{g}: não houve chamamento qualificado", value=False, key=f"{g}__none")
        def _add_g(g=g):
            st.session_state[cnt_key] += 1
            st.session_state.gestor_marcos[g].append(None)
        def _rem_g(g=g):
            if st.session_state[cnt_key] > 1:
                st.session_state[cnt_key] -= 1
                st.session_state.gestor_marcos[g] = st.session_state.gestor_marcos[g][: st.session_state[cnt_key]]
        def _clr_g(g=g):
            st.session_state[cnt_key] = 1
            st.session_state.gestor_marcos[g] = [None]
        if not no_subj:
            for i in range(st.session_state[cnt_key]):
                default_val = st.session_state.gestor_marcos[g][i] or date.today()
                picked = di(f"{g} — data do chamamento #{i+1}", default=default_val, key=f"{g}__marco_{i}")
                st.session_state.gestor_marcos[g][i] = picked
            c1, c2, c3 = st.columns(3)
            c1.button("➕ Adicionar", use_container_width=True, key=f"{g}__add_btn", on_click=_add_g)
            c2.button("➖ Remover última", disabled=st.session_state[cnt_key] <= 1, use_container_width=True, key=f"{g}__rem_btn", on_click=_rem_g)
            c3.button("🗑️ Limpar todas", use_container_width=True, key=f"{g}__clr_btn", on_click=_clr_g)
        else:
            st.session_state[cnt_key] = 1
            st.session_state.gestor_marcos[g] = []

# ======================================================================================
# 3) Funções auxiliares — teste pré-lei e deadline
# ======================================================================================
def _prelaw_consumou_ate_cutoff(ciencia: _date_for_prevcheck, marcos: list[_date_for_prevcheck]) -> bool:
    """Verifica se o quinquênio do regime anterior (ciência) consumou até 18/07/2024,
    considerando apenas marcos entre ciência e cutoff."""
    cutoff = _date_for_prevcheck(2024, 7, 18)
    if not isinstance(ciencia, _date_for_prevcheck):
        return False
    ints_prev = sorted([d for d in marcos if isinstance(d, _date_for_prevcheck) and ciencia <= d <= cutoff])
    start = ciencia
    for d in ints_prev:
        if d >= start:
            start = d
    return start + relativedelta(years=5) <= cutoff

def compute_deadline(data_inicio: date, interrupcoes: list[date], base_anos: int) -> tuple[date, bool]:
    """Retorna (data_final, houve_interrupcao_valida). Ignora marcos anteriores ao termo inicial."""
    ints = sorted([d for d in interrupcoes if d and d >= data_inicio])
    start = data_inicio
    for d in ints:
        if d >= start:
            start = d  # reinicia a contagem a partir do marco
    return start + relativedelta(years=base_anos), (len(ints) > 0)

# ======================================================================================
# 4) Enquadramento intertemporal (GLOBAL — regra consolidada)
# ======================================================================================
fatos_pre_2021 = (termo_material < date(2021, 7, 18))
cutoff = date(2024, 7, 18)

if decisao_transitada_pre_lc == "Sim":
    sugerido = "Fora do alcance: decisão anterior a 18/07/2024"
elif not fatos_pre_2021:
    # Fatos ≥ 18/07/2021 → novo regime (5 anos do fato/cessação), independentemente da ciência/autuação
    sugerido = "Novo regime (art. 5º-A)"
else:
    # Fatos < 18/07/2021 → primeiro TESTE PRÉ-LEI: consumou até 18/07/2024 pelo regime anterior (quinquênio da ciência)?
    if _prelaw_consumou_ate_cutoff(data_ciencia, g_interrupcoes):
        sugerido = "Prescrição consumada antes da lei"
    else:
        # NÃO consumou → Transição bienal (18/07/2024 → 18/07/2026), mesmo que a ciência/autuação seja posterior
        sugerido = "Transição 2 anos (LC 220/24)"

enquadramento = st.selectbox(
    "Selecione o enquadramento (global; ajuste se necessário)",
    [
        "Novo regime (art. 5º-A)",
        "Transição 2 anos (LC 220/24)",
        "Prescrição consumada antes da lei",
        "Fora do alcance: decisão anterior a 18/07/2024",
    ],
    index=[
        "Novo regime (art. 5º-A)",
        "Transição 2 anos (LC 220/24)",
        "Prescrição consumada antes da lei",
        "Fora do alcance: decisão anterior a 18/07/2024",
    ].index(sugerido),
    help=("Chave intertemporal\n"
          "• Fatos < 18/07/2021 → Teste pré-lei: quinquênio da ciência até 18/07/2024; se não consumou, Transição (18/07/2024 → 18/07/2026).\n"
          "• Fatos ≥ 18/07/2021 → Novo regime (5 anos do fato/cessação).\n"
          "• Fora do alcance → decisão adm. transitada até 18/07/2024."),
)

# ======================================================================================
# 5) Prescrição intercorrente (§ 1º)
# ======================================================================================
st.subheader("Prescrição intercorrente (§ 1º)")
st.caption("Paralisação > 3 anos sem julgamento/despacho? Caso positivo, informe as datas.")
check_intercorrente = st.checkbox("Checar intercorrente?", value=False)

data_ultimo_ato = None
idata_subseq = None
if check_intercorrente:
    c1, c2 = st.columns(2)
    with c1:
        data_ultimo_ato = di("Data do último ato útil", default=date.today())
    with c2:
        use_hoje = st.checkbox("Usar a data de hoje como termo final", value=True)
        if use_hoje:
            idata_subseq = date.today()
        else:
            idata_subseq = di("Data do ato subsequente", default=date.today())

# ======================================================================================
# 6) Motor de cálculo por gestor
# ======================================================================================
def calcular_por_gestor(nome_gestor: str,
                        enquadramento: str,
                        termo_material: date,
                        data_ciencia: date,
                        global_marcos: list[date],
                        subj_marcos: list[date],
                        aplicar_prazo_penal: str,
                        prazo_penal_anos: int | None,
                        check_intercorrente: bool,
                        data_ultimo_ato: date | None,
                        idata_subseq: date | None) -> dict:
    resultado = {}

    # Prescrição antes da lei — bloco exclusivo
    if enquadramento == "Prescrição consumada antes da lei":
        cutoff = date(2024, 7, 18)
        ciencia = data_ciencia if isinstance(data_ciencia, date) else None
        ints_prev = [d for d in (global_marcos + subj_marcos) if isinstance(d, date) and d <= cutoff and (ciencia is None or d >= ciencia)]

        def _prelaw_date(ciencia, ints):
            if not ciencia:
                return None
            ints_prev_sorted = sorted(ints)
            start = ciencia
            for d in ints_prev_sorted:
                if d >= start:
                    start = d
            return start + relativedelta(years=5)

        data_prelaw = _prelaw_date(ciencia, ints_prev)
        resultado["sit"] = "Prescrição reconhecida (regime anterior)"
        resultado["detalhe"] = (f"Consumação em {fmt_br(data_prelaw)} (antes de 18/07/2024)."
                                if isinstance(data_prelaw, date) else
                                "Consumação integral antes de 18/07/2024 (regime anterior).")
        resultado["natureza"] = natureza
        resultado["conduta"] = conduta
        resultado["termo_inicial"] = ciencia
        resultado["termo_inicial_label"] = "Ciência (TCE-RJ) — regime anterior"
        resultado["base"] = "quinquenal (regime anterior)"
        resultado["prazo_final"] = data_prelaw
        resultado["interrupcoes"] = sorted(ints_prev)
        return resultado

    # Interrupções consideradas por regime
    if enquadramento == "Transição 2 anos (LC 220/24)":
        interrupcoes = sorted([d for d in (global_marcos + subj_marcos) if isinstance(d, date) and d >= date(2024, 7, 18)])
    elif enquadramento == "Novo regime (art. 5º-A)":
        interrupcoes = sorted([d for d in (global_marcos + subj_marcos) if isinstance(d, date) and d >= termo_material])
    else:
        interrupcoes = sorted([d for d in (global_marcos + subj_marcos) if isinstance(d, date)])

    # Base de prazo (penal prevalece)
    if aplicar_prazo_penal == "Sim" and prazo_penal_anos:
        base_anos = int(prazo_penal_anos)
        base_label = f"prazo penal ({base_anos} anos)"
    else:
        if enquadramento == "Novo regime (art. 5º-A)":
            base_anos = 5; base_label = "quinquenal"
        elif enquadramento == "Transição 2 anos (LC 220/24)":
            base_anos = 2; base_label = "bienal (transição)"
        else:
            base_anos = 5; base_label = "quinquenal"

    # Termo inicial efetivo
    if enquadramento == "Novo regime (art. 5º-A)":
        termo_inicial_efetivo = termo_material
        termo_inicial_label = "Termo inicial (fato/cessação)"
    elif enquadramento == "Transição 2 anos (LC 220/24)":
        termo_inicial_efetivo = date(2024, 7, 18)
        termo_inicial_label = "Transição (18/07/2024)"
    else:
        termo_inicial_efetivo = data_ciencia
        termo_inicial_label = "Ciência (TCE-RJ)"

    prazo_final, has_valid_interruptions = compute_deadline(termo_inicial_efetivo, interrupcoes, base_anos)

    # Intercorrente
    intercorrente = False
    periodo_intercorrente = None
    if check_intercorrente and data_ultimo_ato and idata_subseq:
        dias = (idata_subseq - data_ultimo_ato).days
        if dias >= 365 * 3:
            intercorrente = True
            periodo_intercorrente = dias

    hoje = date.today()
    interrupcoes_consideradas = sorted([d for d in interrupcoes if d and d >= termo_inicial_efetivo])

    if intercorrente:
        resultado["sit"] = "Prescrição intercorrente"
        resultado["detalhe"] = f"Paralisação superior a 3 anos ({periodo_intercorrente} dias)."
    else:
        if hoje >= prazo_final:
            resultado["sit"] = "Prescrição consumada"
            resultado["detalhe"] = f"Esgotado o prazo {base_label}: {fmt_br(prazo_final)}."
        else:
            resultado["sit"] = "Não prescrito"
            resultado["detalhe"] = f"Data-alvo projetada ({base_label}): {fmt_br(prazo_final)}."

    resultado["natureza"] = natureza
    resultado["conduta"] = conduta
    resultado["termo_inicial"] = termo_inicial_efetivo
    resultado["termo_inicial_label"] = termo_inicial_label
    resultado["prazo_final"] = prazo_final
    resultado["base"] = base_label
    resultado["interrupcoes"] = interrupcoes_consideradas
    return resultado

# ======================================================================================
# 7) Resultados por gestor (cards) + linhas para Excel
# ======================================================================================
st.markdown("### Resultados por gestor")

def _color_for_status(s: str) -> str:
    s = (s or '').lower()
    if 'prescrição consumada' in s or 'intercorrente' in s or 'prescrição reconhecida' in s:
        return '#D93025'
    elif 'não prescrito' in s:
        return '#1E8E3E'
    else:
        return '#1A73E8'

rows_resumo: list[dict] = []

# Partição de marcos gerais para referência (pré/pós-lei)
g_interrupcoes_pre = sorted([d for d in g_interrupcoes if d <= date(2024, 7, 18)])
g_interrupcoes_pos = sorted([d for d in g_interrupcoes if d >= date(2024, 7, 18)])

for g in gestores:
    subj_list = [d for d in st.session_state.gestor_marcos.get(g, []) if isinstance(d, date)]
    res = calcular_por_gestor(
        nome_gestor=g,
        enquadramento=enquadramento,
        termo_material=termo_material,
        data_ciencia=data_ciencia,
        global_marcos=g_interrupcoes,
        subj_marcos=subj_list,
        aplicar_prazo_penal=aplicar_prazo_penal,
        prazo_penal_anos=prazo_penal_anos,
        check_intercorrente=check_intercorrente,
        data_ultimo_ato=data_ultimo_ato,
        idata_subseq=idata_subseq
    )

    _sit = res.get('sit', '—')
    _status_color = _color_for_status(_sit)
    _termo = res.get('termo_inicial')
    _prazo = res.get('prazo_final')
    _ints = res.get('interrupcoes', [])
    _ints_str = ", ".join([fmt_br(d) for d in _ints]) if _ints else '—'

    _html = f"""
    <div style='border:1px solid {_status_color}; padding:16px; border-radius:12px; margin-bottom:8px;'>
      <div style='font-weight:700; font-size:1.05rem; color:{_status_color};'>[{g}] Situação: {res.get('sit','—')}</div>
      <div style='margin-top:6px;'>{res.get('detalhe','—')}</div>
      <hr style='border:none; border-top:1px dashed #ddd; margin:12px 0;'>
      <div style='display:grid; grid-template-columns: 1fr 1fr; gap:8px;'>
        <div><b>Enquadramento:</b> {enquadramento}</div>
        <div><b>Base:</b> {res.get('base','—')}</div>
        <div><b>Natureza:</b> {res.get('natureza','—')}</div>
        <div><b>Conduta:</b> {res.get('conduta','—')}</div>
        <div><b>Termo inicial (cálculo):</b> {(fmt_br(_termo) if isinstance(_termo, date) else '—')} ({res.get('termo_inicial_label','')})</div>
        <div><b>Data-alvo de prescrição:</b> {(fmt_br(_prazo) if isinstance(_prazo, date) else '—')}</div>
        <div><b>Ciência considerada (TCE-RJ):</b> {fmt_br(data_ciencia)}</div>
        <div><b>Data do fato/cessação:</b> {fmt_br(termo_material)}</div>
        <div style='grid-column: 1 / -1;'><b>Interrupções consideradas:</b> {_ints_str}</div>
      </div>
    </div>
    """
    st.markdown(_html, unsafe_allow_html=True)

    rows_resumo.append({
        "Gestor": g,
        "Enquadramento": enquadramento,
        "Situação": res.get('sit','—'),
        "Base": res.get('base','—'),
        "Termo material": fmt_br(termo_material),
        "Ciência (TCE-RJ)": fmt_br(data_ciencia),
        "Termo inicial efetivo": fmt_br(_termo) if isinstance(_termo, date) else '',
        "Data atual de prescrição": fmt_br(_prazo) if isinstance(_prazo, date) else '',
        "Interrupções consideradas": ", ".join([fmt_br(d) for d in _ints]) if _ints else ''
    })

# ======================================================================================
# 8) Coletas auxiliares para exportação
# ======================================================================================
rows_marcos_gerais = [{"marco_geral_data": fmt_br(d)} for d in sorted(g_interrupcoes)]
rows_marcos_subj = []
for g in gestores:
    cham_g = [d for d in st.session_state.gestor_marcos.get(g, []) if isinstance(d, date)]
    for d in cham_g:
        rows_marcos_subj.append({"gestor": g, "chamamento_data": fmt_br(d)})

parametros_do_caso = {
    "natureza": natureza,
    "conduta": conduta,
    "data_autuacao": fmt_br(data_autuacao),
    "data_ciencia": fmt_br(data_ciencia),
    "termo_material_label": termo_material_label,
    "termo_material_data": fmt_br(termo_material),
    "decisao_transitada_pre_lc_220_2024": decisao_transitada_pre_lc,
    "aplicar_prazo_penal": aplicar_prazo_penal,
    "prazo_penal_anos": int(prazo_penal_anos) if prazo_penal_anos else "",
    "enquadramento_global": enquadramento,
    "check_intercorrente": "Sim" if check_intercorrente else "Não",
    "intercorrente_ultimo_ato": fmt_br(data_ultimo_ato) if data_ultimo_ato else "",
    "intercorrente_ato_subseq_ou_hoje": fmt_br(idata_subseq) if idata_subseq else "",
}

por_gestor_details = {}
for g in gestores:
    cham_g = [d for d in st.session_state.gestor_marcos.get(g, []) if isinstance(d, date)]
    linhas = [
        {"campo": "Gestor", "valor": g},
        {"campo": "Enquadramento (global)", "valor": enquadramento},
        {"campo": "Natureza", "valor": natureza},
        {"campo": "Conduta", "valor": conduta},
        {"campo": "Termo material — label", "valor": termo_material_label},
        {"campo": "Termo material — data", "valor": fmt_br(termo_material)},
        {"campo": "Ciência considerada (TCE-RJ)", "valor": fmt_br(data_ciencia)},
        {"campo": "Marcos gerais (todas as datas)", "valor": ", ".join(sorted({fmt_br(d) for d in g_interrupcoes})) if g_interrupcoes else "—"},
        {"campo": f"Chamamentos qualificados — {g}", "valor": ", ".join(sorted({fmt_br(d) for d in cham_g})) if cham_g else "—"},
    ]
    por_gestor_details[g] = {"linhas": linhas}

# ======================================================================================
# 9) Exportação — botão Excel
# ======================================================================================
st.markdown("#### Exportação (Excel)")
if rows_resumo:
    xlsx_bytes = make_excel_bytes_expanded(
        rows_resumo=rows_resumo,
        rows_marcos_gerais=rows_marcos_gerais,
        rows_marcos_subj=rows_marcos_subj,
        parametros=parametros_do_caso,
        por_gestor_details=por_gestor_details
    )
    st.download_button(
        "⬇️ Baixar resumo (Excel)",
        data=xlsx_bytes,
        file_name=f"prescricao_resultados_gestores_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
else:
    st.info("Preencha os dados e calcule ao menos um gestor para habilitar a exportação.")
